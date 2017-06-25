from time import strftime, gmtime
import time
from openpyxl import load_workbook
import pandas as pd
from conf import EXCEL_FILE


def iniatilize_results_file(measures):
    writer = pd.ExcelWriter(EXCEL_FILE, engine='openpyxl')
    columns = ['SCREEN_NAME'] + measures
    initialization_list = ['0'] * columns.__len__()
    pd.DataFrame([initialization_list]).to_excel(writer, header=columns, index=False)
    writer.save()
    return writer


def write_user_data_to_file(index, writer, user, user_dict_values):
    wb = load_workbook(EXCEL_FILE, read_only=True)
    pd.DataFrame([[user.name] + user_dict_values]).to_excel(writer, startrow=index, index=False, header=False)
    writer.save()


def write_to_excel(self):
    EXCEL_FILE = 'ron.xlsx'
    writer = pd.ExcelWriter(EXCEL_FILE, engine='openpyxl')
    pd.DataFrame().to_excel(writer)
    writer.save()
    i = 0
    while i < 10:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        if i == 0:
            pd.DataFrame([['a', 'b', 'c']]).to_excel(writer, header=['ron', 'moses', 'arik'], index=False)
        else:
            pd.DataFrame([['d', 'e', 'f']]).to_excel(writer, startrow=i + 1, index=False, header=False)
        writer.save()
        i += 1


def get_current_time():
    return str(strftime("%H:%M:%S", gmtime()))


def handle_exception(e, screen_name, count_users):
    if e.__class__.__name__ is 'TweepError':
        print 'Reason: '+e.response.reason+'\nStatus Code: '+str(e.response.status_code)
        if e.response.status_code==429 or e.__class__.__name__ is 'RateLimitError':
            print 'Limit Reached - User: ' + screen_name + ' - User Number:' + str(count_users)
            i = 0
            while i <= 16:
                time.sleep(60)
                i += 1
                print str(i) + ' min passed'
    else:
        print 'EXCEPTION!!!!! - User: ' + screen_name + ' - User Number:' + str(count_users)